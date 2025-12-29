from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.graphics.barcode import code39
from reportlab.lib.utils import ImageReader
import openpyxl
from openpyxl.styles import PatternFill  # Importamos herramienta de pintura
import os


# ******************************************** zona de configuracion ********************************************

class Config:
    """configuracion centralizada del generador de etiquetas"""
    
    # archivo excel
    NOMBRE_EXCEL = "LIBROS FIMGM.xlsx"
    NOMBRE_EXCEL_SALIDA = "LIBROS FIMGM_PINTADO.xlsx" # nombre del archivo coloreado
    
    FILA_INICIAL = 3  # configuracion de fila inicial
    COLUMNA_CODIGOS = "A"        # codigo de barras
    COLUMNA_ESTANTERIA = "N"     # ubicacion en la estanteria
    
    ABREVIACION_FACULTAD = "FIMGM"
    
    # fuentes
    RUTA_FUENTE = "OpenSans-Bold.ttf"
    RUTA_FUENTE_CODE = "OpenSans-Semibold.ttf"
    
    # configuracion de imagenes - logos
    RUTA_LOGO_UNASAM = "logo-unasam.png"
    RUTA_LOGO_FACULTAD = "facultad.jpg"
    
    # dimensiones y ubicacion de imagenes
    ALTO_IMAGENES = 0.7 * cm
    
    # margenes horizontales para logos
    MARGEN_X_LOGO_UNASAM = 0.7 * cm      # margen izquierdo para logo unasam
    MARGEN_X_LOGO_FACULTAD = 0.7 * cm    # margen derecho para logo facultad
    
    # distancia vertical de las imagenes desde el codigo
    DISTANCIA_Y_DESDE_CODIGO = -1.3 * cm
    
    # margen de pagina
    MARGEN_SUPERIOR = 1.1 * cm
    MARGEN_IZQUIERDO = 0.4 * cm
    
    # --- grid de etiquetas ---
    FILAS = 8
    COLUMNAS = 3
    CUADROS_POR_HOJA = FILAS * COLUMNAS
    
    ESPACIO_HORIZONTAL = 0.15 * cm  # espacio entre columnas
    ESPACIO_VERTICAL = 0.06 * cm    # espacio entre filas
    Y_INICIAL_GRID = 2.70 * cm
    
    # dimensiones de cuadro individual 
    ANCHO_CUADRO = 6.56 * cm
    ALTO_CUADRO = 3.28 * cm
    
    # tipografia 
    TAMANO_FUENTE_TITULO = 13
    TAMANO_FUENTE_CUADRO = 10
    TAMANO_FUENTE_CODIGO = 8
    
    # codigo de barras visual 
    ANCHO_BARRAS = 0.05 * cm
    ALTO_BARRAS = 0.94 * cm
    MARGEN_HORIZONTAL_BARRAS = 0.1 * cm
    
    # codigo de barras textual
    MARGEN_HORIZONTAL_TEXTO = 1 * cm 
    SEPARACION_TEXTO_BARRAS = 0.3 * cm
    
    # ajustes finos
    AJUSTE_VERTICAL_CODIGO = 0.1 * cm  # ajuste vertical del bloque barras + texto
    
    # contenido del titulo
    TITULO_CUADRO = "Sistema de Bibliotecas UNASAM"


# ********************************************** lectura de datos **********************************************

class LectorExcel:
    """maneja la lectura del archivo excel"""
    
    def __init__(self, config):
        self.config = config
        self.workbook = None
        self.sheet = None
    
    def cargar_excel(self):
        """carga el archivo excel y retorna true si fue exitoso"""
        try:
            # data_only=True obtiene los valores calculados, no las formulas
            self.workbook = openpyxl.load_workbook(
                self.config.NOMBRE_EXCEL, 
                data_only=True
            )
            self.sheet = self.workbook.active
            print(f"excel cargado (lectura) - {self.config.NOMBRE_EXCEL}")
            return True
        except FileNotFoundError:
            print(f"error - no se encontro '{self.config.NOMBRE_EXCEL}'")
            return False
        except Exception as e:
            print(f"error al cargar excel - {e}")
            return False
    
    def obtener_ultima_fila(self):
        """obtiene la ultima fila con datos en la columna de estanteria"""
        ultima_fila = self.sheet.max_row
        
        # buscar la ultima fila que tenga datos en columna estanteria
        for fila in range(ultima_fila, 0, -1):
            celda = f"{self.config.COLUMNA_ESTANTERIA}{fila}"
            valor = self.sheet[celda].value
            if valor is not None and str(valor).strip() != "":
                return fila
        
        return self.config.FILA_INICIAL
    
    def leer_valor_estanteria(self, fila):
        """lee el valor de estanteria de una fila especifica"""
        celda = f"{self.config.COLUMNA_ESTANTERIA}{fila}"
        valor = self.sheet[celda].value
        return str(valor).strip() if valor else ""
    
    def leer_codigos_rango(self, fila_inicio, fila_fin):
        """lee los codigos de barras de un rango especifico"""
        codigos = []
        for fila in range(fila_inicio, fila_fin + 1):
            celda = f"{self.config.COLUMNA_CODIGOS}{fila}"
            valor = self.sheet[celda].value
            
            if valor is None or str(valor).strip() == "":
                codigos.append("*0*")
            else:
                codigos.append(str(valor).strip())
        
        return codigos
    
    def cerrar(self):
        """cierra el archivo excel"""
        if self.workbook:
            self.workbook.close()


# ********************************************** pintor de excel **********************************************

class PintorExcel:
    """maneja el pintado de celdas en el excel"""
    
    def __init__(self, config):
        self.config = config
        self.workbook = None
        self.sheet = None
        # Definimos una paleta de colores suaves (hexadecimal ARGB)
        self.colores = [
            PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid"), # Verde Claro
            PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"), # Azul Claro
            PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid"), # Amarillo Claro
            PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"), # Naranja Claro
        ]
        
    def cargar_para_pintar(self):
        """carga el excel en modo escritura"""
        try:
            # Cargamos SIN data_only para intentar preservar el formato original
            self.workbook = openpyxl.load_workbook(self.config.NOMBRE_EXCEL)
            self.sheet = self.workbook.active
            print(f"excel cargado (escritura) - listo para pintar")
            return True
        except Exception as e:
            print(f"error al cargar excel para pintar - {e}")
            return False
            
    def pintar_rango(self, fila_inicio, fila_fin, indice_lote):
        """pinta las celdas de estanteria para un rango especifico"""
        # Seleccionar color basado en el indice del lote (rotativo)
        color_actual = self.colores[indice_lote % len(self.colores)]
        columna = self.config.COLUMNA_ESTANTERIA
        
        for fila in range(fila_inicio, fila_fin + 1):
            celda = f"{columna}{fila}"
            self.sheet[celda].fill = color_actual
            
    def guardar(self):
        """guarda el archivo modificado"""
        try:
            self.workbook.save(self.config.NOMBRE_EXCEL_SALIDA)
            print(f"✓ Excel pintado guardado como: {self.config.NOMBRE_EXCEL_SALIDA}")
            self.workbook.close()
            return True
        except Exception as e:
            print(f"error al guardar excel pintado - {e}")
            return False


# ********************************************** procesador de lotes **********************************************

class ProcesadorLotes:
    """maneja la logica de division automatica en lotes de 72 filas"""
    
    def __init__(self, lector, config):
        self.lector = lector
        self.config = config
        self.ultima_fila_excel = lector.obtener_ultima_fila()
    
    def calcular_lotes(self):
        """calcula todos los lotes que se deben generar"""
        lotes = []
        fila_actual = self.config.FILA_INICIAL
        
        print(f"\n{'=' * 60}")
        print(f"calculando lotes desde fila {fila_actual} hasta {self.ultima_fila_excel}")
        print(f"{'=' * 60}\n")
        
        while fila_actual <= self.ultima_fila_excel:
            lote = self._calcular_lote_individual(fila_actual)
            
            if lote:
                lotes.append(lote)
                print(f"lote {len(lotes)}: filas {lote['fila_inicio']}-{lote['fila_fin']} "
                      f"({lote['total_filas']} filas) - rango [{lote['rango_inicial']} - {lote['rango_final']}]")
                fila_actual = lote['fila_fin'] + 1
            else:
                break
        
        print(f"\n{'=' * 60}")
        print(f"total de lotes calculados: {len(lotes)}")
        print(f"{'=' * 60}\n")
        
        return lotes
    
    def _calcular_lote_individual(self, fila_inicio):
        """calcula un lote individual desde una fila inicial"""
        # verificar que no excedamos el excel
        if fila_inicio > self.ultima_fila_excel:
            return None
        
        # caso especial - quedan menos de 72 filas
        filas_restantes = self.ultima_fila_excel - fila_inicio + 1
        if filas_restantes <= 72:
            return {
                'fila_inicio': fila_inicio,
                'fila_fin': self.ultima_fila_excel,
                'total_filas': filas_restantes,
                'rango_inicial': self.lector.leer_valor_estanteria(fila_inicio),
                'rango_final': self.lector.leer_valor_estanteria(self.ultima_fila_excel)
            }
        
        # tomar 72 filas tentativamente
        fila_fin_tentativa = fila_inicio + 71
        
        # verificar continuidad
        valor_72 = self.lector.leer_valor_estanteria(fila_fin_tentativa)
        valor_73 = self.lector.leer_valor_estanteria(fila_fin_tentativa + 1)
        
        # **** caso 1 - no hay continuidad - lote normal de 72
        if valor_72 != valor_73:
            return {
                'fila_inicio': fila_inicio,
                'fila_fin': fila_fin_tentativa,
                'total_filas': 72,
                'rango_inicial': self.lector.leer_valor_estanteria(fila_inicio),
                'rango_final': valor_72
            }
        
        # **** caso 2 - hay continuidad - buscar inicio y fin del grupo
        inicio_grupo = self._encontrar_inicio_grupo(fila_fin_tentativa, valor_72)
        fin_grupo = self._encontrar_fin_grupo(fila_fin_tentativa + 1, valor_72)
        
        total_filas_grupo = fin_grupo - inicio_grupo + 1
        
        # **** caso 2a: mega-grupo - mas de 72 filas
        if total_filas_grupo > 72:
            # verificar si hay filas antes del mega-grupo
            if inicio_grupo > fila_inicio:
                # generar lote con filas antes del mega-grupo
                return {
                    'fila_inicio': fila_inicio,
                    'fila_fin': inicio_grupo - 1,
                    'total_filas': inicio_grupo - fila_inicio,
                    'rango_inicial': self.lector.leer_valor_estanteria(fila_inicio),
                    'rango_final': self.lector.leer_valor_estanteria(inicio_grupo - 1)
                }
            else:
                # *** todo el lote es el mega-grupo
                return {
                    'fila_inicio': inicio_grupo,
                    'fila_fin': fin_grupo,
                    'total_filas': total_filas_grupo,
                    'rango_inicial': valor_72,
                    'rango_final': valor_72,
                    'es_mega_grupo': True
                }
        
        # **** caso 2b: grupo normal con desbordamiento - menos de 72 filas
        # excluir el grupo del lote actual
        return {
            'fila_inicio': fila_inicio,
            'fila_fin': inicio_grupo - 1,
            'total_filas': inicio_grupo - fila_inicio,
            'rango_inicial': self.lector.leer_valor_estanteria(fila_inicio),
            'rango_final': self.lector.leer_valor_estanteria(inicio_grupo - 1)
        }
    
    def _encontrar_inicio_grupo(self, fila_desde, valor_buscado):
        """retrocede para encontrar donde empieza el grupo con el mismo valor"""
        fila = fila_desde
        while fila >= self.config.FILA_INICIAL:
            valor_actual = self.lector.leer_valor_estanteria(fila)
            if valor_actual != valor_buscado:
                return fila + 1
            fila -= 1
        return self.config.FILA_INICIAL
    
    def _encontrar_fin_grupo(self, fila_desde, valor_buscado):
        """avanza para encontrar donde termina el grupo con el mismo valor"""
        fila = fila_desde
        while fila <= self.ultima_fila_excel:
            valor_actual = self.lector.leer_valor_estanteria(fila)
            if valor_actual != valor_buscado:
                return fila - 1
            fila += 1
        return self.ultima_fila_excel


# ********************************************** generacion del pdf **********************************************

class GeneradorEtiquetas:
    """genera el pdf con las etiquetas de codigos de barras"""
    
    def __init__(self, config):
        self.config = config
        self.fuente_bold = None
        self.fuente_code = None
        self._cargar_fuentes()
    
    # inicializacion 
    
    def _cargar_fuentes(self):
        """carga las fuentes personalizadas o usa alternativas"""
        # fuente bold - titulos
        if os.path.exists(self.config.RUTA_FUENTE):
            pdfmetrics.registerFont(TTFont('OpenSans-Bold', self.config.RUTA_FUENTE))
            self.fuente_bold = "OpenSans-Bold"
        else:
            print(f"aviso - no se encontro '{self.config.RUTA_FUENTE}', usando helvetica-bold")
            self.fuente_bold = "Helvetica-Bold"
        
        # fuente code - texto del codigo
        if os.path.exists(self.config.RUTA_FUENTE_CODE):
            pdfmetrics.registerFont(TTFont('OpenSans-Code', self.config.RUTA_FUENTE_CODE))
            self.fuente_code = "OpenSans-Code"
        else:
            print(f"aviso - no se encontro '{self.config.RUTA_FUENTE_CODE}', usando fuente principal")
            self.fuente_code = self.fuente_bold

    def _calcular_siguiente_numero(self):
        """busca el siguiente numero de archivo basado en lo que existe en la carpeta"""
        facultad = self.config.ABREVIACION_FACULTAD
        archivos = [f for f in os.listdir('.') if f.endswith('.pdf')]
        
        max_numero = 0
        
        for archivo in archivos:
            if facultad in archivo:
                partes = archivo.split(facultad, 1)
                
                if len(partes) > 1:
                    prefijo = partes[0]
                    
                    if prefijo.isdigit():
                        numero = int(prefijo)
                        if numero > max_numero:
                            max_numero = numero
        
        siguiente = max_numero + 1
        return str(siguiente)
    
    def _obtener_nombre_archivo(self, numero_archivo, rango_inicial, rango_final):
        """genera el nombre del archivo pdf"""
        rango_inicial_limpio = rango_inicial.replace('*', '').replace('/', '-').replace('\\', '-').replace(':', '-')
        rango_final_limpio = rango_final.replace('*', '').replace('/', '-').replace('\\', '-').replace(':', '-')
        
        return f"{numero_archivo}{self.config.ABREVIACION_FACULTAD} {rango_inicial_limpio} - {rango_final_limpio}.pdf"
    
    # **************************** dibujo de titulos ****************************
    
    def _dibujar_titulo_principal(self, c, ancho_hoja, alto_hoja, rango_inicial, rango_final):
        """dibuja el titulo principal en la parte superior de la pagina"""
        pos_x = 5.52 * cm
        ancho = 9.97 * cm
        alto = 1.18 * cm
        pos_y = alto_hoja - self.config.MARGEN_SUPERIOR - alto
        centro_x = pos_x + (ancho / 2)
        
        c.setFillColorRGB(0, 0, 0)
        c.setFont(self.fuente_bold, self.config.TAMANO_FUENTE_TITULO)
        
        # linea 1 - biblioteca y ubicacion
        y_linea1 = pos_y + alto - 0.35 * cm
        c.drawCentredString(
            centro_x, 
            y_linea1, 
            f"BIBLIOTECA {self.config.ABREVIACION_FACULTAD} - UBICACION ESTANTERIA"
        )
        
        # linea 2 - rango de estanteria
        y_linea2 = pos_y + 0.25 * cm
        c.drawCentredString(
            centro_x, 
            y_linea2, 
            f"{rango_inicial} - {rango_final}"
        )
    
    # **************************** dibujo de elementos - imagenes ****************************
    
    def _dibujar_imagen(self, c, ruta_imagen, x, y, alto_deseado):
        """dibuja una imagen redimensionada proporcionalmente"""
        if not os.path.exists(ruta_imagen):
            return 0

        try:
            img_utils = ImageReader(ruta_imagen)
            ancho_real, alto_real = img_utils.getSize()
            aspect_ratio = ancho_real / alto_real
            nuevo_ancho = alto_deseado * aspect_ratio
            c.drawImage(ruta_imagen, x, y, width=nuevo_ancho, height=alto_deseado, mask='auto')
            return nuevo_ancho
        except Exception as e:
            print(f"error al dibujar imagen {ruta_imagen} - {e}")
            return 0
            
    # **************************** dibujo de elementos - codigo de barras ****************************
    
    def _dibujar_codigo_barras(self, c, x_cuadro, y_base, codigo):
        """dibuja el codigo de barras visual - barras negras"""
        codigo_limpio = codigo.replace("*", "")
        ancho_maximo = self.config.ANCHO_CUADRO - (2 * self.config.MARGEN_HORIZONTAL_BARRAS)
        
        barcode = code39.Standard39(
            codigo_limpio,
            barHeight=self.config.ALTO_BARRAS,
            barWidth=self.config.ANCHO_BARRAS,
            checksum=0,
            humanReadable=False
        )
        
        if barcode.width > ancho_maximo:
            factor_reduccion = ancho_maximo / barcode.width
            nuevo_ancho_barra = self.config.ANCHO_BARRAS * factor_reduccion
            barcode = code39.Standard39(
                codigo_limpio,
                barHeight=self.config.ALTO_BARRAS,
                barWidth=nuevo_ancho_barra,
                checksum=0,
                humanReadable=False
            )
        
        centro_x_cuadro = x_cuadro + (self.config.ANCHO_CUADRO / 2)
        x_barcode = centro_x_cuadro - (barcode.width / 2)
        
        barcode.drawOn(c, x_barcode, y_base)
        return barcode.width
    
    def _dibujar_texto_codigo(self, c, x_cuadro, y_base, codigo):
        """dibuja el codigo en formato textual con justificacion expandida"""
        margen = self.config.MARGEN_HORIZONTAL_TEXTO 
        ancho_util_texto = self.config.ANCHO_CUADRO - (2 * margen)
        x_inicio_texto = x_cuadro + margen
        
        c.setFont(self.fuente_code, self.config.TAMANO_FUENTE_CODIGO)
        
        ancho_texto_puro = c.stringWidth(codigo, self.fuente_code, self.config.TAMANO_FUENTE_CODIGO)
        tamano_actual = self.config.TAMANO_FUENTE_CODIGO
        
        if ancho_texto_puro > ancho_util_texto:
            factor = ancho_util_texto / ancho_texto_puro
            tamano_actual = tamano_actual * factor
            c.setFont(self.fuente_code, tamano_actual)
        
        num_caracteres = len(codigo)
        
        if num_caracteres <= 1:
            c.drawCentredString(x_inicio_texto + (ancho_util_texto / 2), y_base, codigo)
            return
        
        ancho_solo_letras = 0
        anchos_individuales = []
        for letra in codigo:
            w = c.stringWidth(letra, self.fuente_code, tamano_actual)
            anchos_individuales.append(w)
            ancho_solo_letras += w
        
        espacio_sobrante = ancho_util_texto - ancho_solo_letras
        if espacio_sobrante < 0:
            espacio_sobrante = 0
        
        gap = espacio_sobrante / (num_caracteres - 1)
        
        x_cursor = x_inicio_texto
        for i, letra in enumerate(codigo):
            c.drawString(x_cursor, y_base, letra)
            x_cursor += anchos_individuales[i] + gap
    
    # ************************************** dibujo de elementos - cuadro individual completo **************************************
    
    def _dibujar_cuadro(self, c, x, y, codigo):
        """dibuja un cuadro individual con titulo, codigo de barras y texto"""
        c.setLineWidth(1)
        c.setStrokeColorRGB(0, 0, 0)
        c.setFillColorRGB(0, 0, 0)
        c.rect(x, y, self.config.ANCHO_CUADRO, self.config.ALTO_CUADRO)
        
        centro_x = x + (self.config.ANCHO_CUADRO / 2)
        
        c.setFont(self.fuente_bold, self.config.TAMANO_FUENTE_CUADRO)
        alto_titulo = 0.4 * cm
        y_titulo = y + self.config.ALTO_CUADRO - alto_titulo - 0.2 * cm
        c.drawCentredString(centro_x, y_titulo, self.config.TITULO_CUADRO)
        
        altura_total_visual = 1.34 * cm
        espacio_texto_total = altura_total_visual - self.config.ALTO_BARRAS
        
        y_base_bloque = y + (self.config.ALTO_CUADRO - altura_total_visual) / 2
        y_base_bloque += self.config.AJUSTE_VERTICAL_CODIGO
        
        y_barras = y_base_bloque + espacio_texto_total + 0.03 * cm
        y_texto = y_barras - self.config.SEPARACION_TEXTO_BARRAS
        y_imagenes = y_barras + self.config.DISTANCIA_Y_DESDE_CODIGO
        
        x_logo_unasam = x + self.config.MARGEN_X_LOGO_UNASAM
        self._dibujar_imagen(c, self.config.RUTA_LOGO_UNASAM, x_logo_unasam, y_imagenes, self.config.ALTO_IMAGENES)
        
        ancho_img_facultad = 0
        if os.path.exists(self.config.RUTA_LOGO_FACULTAD):
            try:
                img_fac = ImageReader(self.config.RUTA_LOGO_FACULTAD)
                w_f, h_f = img_fac.getSize()
                aspect_f = w_f / h_f
                ancho_img_facultad = self.config.ALTO_IMAGENES * aspect_f
            except:
                ancho_img_facultad = 0
        
        if ancho_img_facultad > 0:
            x_logo_facultad = (x + self.config.ANCHO_CUADRO) - self.config.MARGEN_X_LOGO_FACULTAD - ancho_img_facultad
            self._dibujar_imagen(c, self.config.RUTA_LOGO_FACULTAD, x_logo_facultad, y_imagenes, self.config.ALTO_IMAGENES)
        
        self._dibujar_codigo_barras(c, x, y_barras, codigo)
        self._dibujar_texto_codigo(c, x, y_texto, codigo)
    
    # *************************************** calculo de posiciones ***************************************
    
    def _calcular_posiciones_x(self):
        """calcula las posiciones x de las columnas del grid"""
        posiciones = []
        x_actual = self.config.MARGEN_IZQUIERDO
        
        for _ in range(self.config.COLUMNAS):
            posiciones.append(x_actual)
            x_actual += self.config.ANCHO_CUADRO + self.config.ESPACIO_HORIZONTAL
        
        return posiciones
    
    def _calcular_posiciones_y(self, alto_hoja):
        """calcula las posiciones y de las filas del grid"""
        posiciones = []
        y_actual = self.config.Y_INICIAL_GRID
        
        for _ in range(self.config.FILAS):
            y_real = alto_hoja - y_actual - self.config.ALTO_CUADRO
            posiciones.append(y_real)
            y_actual += self.config.ALTO_CUADRO + self.config.ESPACIO_VERTICAL
        
        return posiciones
    
    # ******************************** composicion de pagina ********************************
    
    def _dibujar_pagina(self, c, codigos_pagina, ancho_hoja, alto_hoja, rango_inicial, rango_final):
        """dibuja una pagina completa con titulo y grid de etiquetas"""
        self._dibujar_titulo_principal(c, ancho_hoja, alto_hoja, rango_inicial, rango_final)
        
        posiciones_x = self._calcular_posiciones_x()
        posiciones_y = self._calcular_posiciones_y(alto_hoja)
        
        indice = 0
        for y in posiciones_y:
            for x in posiciones_x:
                if indice < len(codigos_pagina):
                    self._dibujar_cuadro(c, x, y, codigos_pagina[indice])
                    indice += 1
                else:
                    break
            if indice >= len(codigos_pagina):
                break
    
    # **************************** generacion principal ****************************
    
    def generar_pdf_lote(self, lote, codigos, numero_archivo):
        """genera un archivo pdf para un lote especifico"""
        nombre_archivo = self._obtener_nombre_archivo(numero_archivo, lote['rango_inicial'], lote['rango_final'])
        c = canvas.Canvas(nombre_archivo, pagesize=A4)
        ancho_hoja, alto_hoja = A4
        
        total_codigos = len(codigos)
        total_paginas = (total_codigos + self.config.CUADROS_POR_HOJA - 1) // self.config.CUADROS_POR_HOJA
        
        print(f"\ngenerando: {nombre_archivo}")
        print(f"  filas: {lote['fila_inicio']}-{lote['fila_fin']} ({lote['total_filas']} etiquetas)")
        print(f"  paginas: {total_paginas}")
        
        for num_pagina in range(total_paginas):
            inicio = num_pagina * self.config.CUADROS_POR_HOJA
            fin = min(inicio + self.config.CUADROS_POR_HOJA, total_codigos)
            codigos_pagina = codigos[inicio:fin]
            self._dibujar_pagina(c, codigos_pagina, ancho_hoja, alto_hoja, lote['rango_inicial'], lote['rango_final'])
            c.showPage()
        
        c.save()
        print(f"  ✓ generado correctamente")


# ************************************* ejecucion principal *************************************

def main():
    """funcion principal que ejecuta todo el proceso automatizado"""
    config = Config()
    
    # 1. Leer el excel para obtener datos
    lector = LectorExcel(config)
    if not lector.cargar_excel(): 
        return
    
    # 2. Inicializar el pintor (cargar excel para escribir)
    pintor = PintorExcel(config)
    pintor_activo = pintor.cargar_para_pintar()
    
    # 3. Calcular lotes
    procesador = ProcesadorLotes(lector, config)
    lotes = procesador.calcular_lotes()
    
    if not lotes:
        print("error - no se calcularon lotes")
        lector.cerrar()
        return
    
    # 4. Generar PDFs y Pintar Excel
    generador = GeneradorEtiquetas(config)
    numero_inicial = int(generador._calcular_siguiente_numero())
    
    print(f"\n{'=' * 60}")
    print(f"iniciando generacion de {len(lotes)} archivo(s) PDF y pintado de Excel")
    print(f"numero inicial: {numero_inicial}")
    print(f"{'=' * 60}")
    
    for i, lote in enumerate(lotes):
        # Generar PDF
        numero_archivo = str(numero_inicial + i)
        codigos = lector.leer_codigos_rango(lote['fila_inicio'], lote['fila_fin'])
        generador.generar_pdf_lote(lote, codigos, numero_archivo)
        
        # Pintar Excel (si se cargó correctamente)
        if pintor_activo:
            pintor.pintar_rango(lote['fila_inicio'], lote['fila_fin'], i)
    
    # Cerrar lector
    lector.cerrar()
    
    # 5. Guardar el Excel pintado
    if pintor_activo:
        pintor.guardar()
    
    print(f"\n{'=' * 60}")
    print(f"✓ proceso completado exitosamente")
    print(f"  archivos generados: {len(lotes)}")
    if pintor_activo:
        print(f"  excel pintado: {config.NOMBRE_EXCEL_SALIDA}")
    print(f"{'=' * 60}\n")


if __name__ == "__main__":
    main()